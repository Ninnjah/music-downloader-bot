import logging
import re
from datetime import datetime
from io import BytesIO
from operator import itemgetter

from aiogram import Router
from aiogram.types import BufferedInputFile, CallbackQuery, Message
from aiogram_dialog import Dialog, DialogManager, Window
from aiogram_dialog.widgets.input import ManagedTextInput, TextInput
from aiogram_dialog.widgets.kbd import (
    Button,
    Cancel,
    Checkbox,
    ManagedCheckbox,
    ScrollingGroup,
    Select,
)
from aiogram_dialog.widgets.text import Format
from openpyxl import Workbook

from tgbot.handlers.admin.states.users import ProfileUserSG, UserSG
from tgbot.services.l10n_dialog import L10NFormat
from tgbot.services.repository import Repo

logger = logging.getLogger(__name__)
router = Router(name=__name__)
# TODO FULL REWRITE
# TODO RENAME TO USER_LIST
# TODO ADD EXCEL HELPER


def process_user_id(query: str) -> int:
    query = re.search(r"\d+", query)
    if query is None:
        raise ValueError

    return int(query.group(0))


async def get_user_id(dialog_manager: DialogManager, **kwargs):
    user_id = dialog_manager.dialog_data.get("user_id")
    return {"user_id": str(user_id)}


async def get_user(dialog_manager: DialogManager, repo: Repo, **kwargs):
    user_id = dialog_manager.dialog_data.get("user_id") or dialog_manager.start_data.get("user_id")
    dialog_manager.dialog_data["user_id"] = user_id
    user = await repo.get_user(user_id)

    ban_checkbox = dialog_manager.find("user_ban_ch")
    if ban_checkbox:
        await ban_checkbox.set_checked(user.banned)

    return {
        "user_id": user.id,
        "firstname": user.firstname,
        "username": user.username or 0,
        "banned": user.banned,
        "created_on": user.created_on.strftime("%Y.%m.%d %H:%M"),
        "updated_on": user.updated_on.strftime("%Y.%m.%d %H:%M"),
    }


async def get_users(dialog_manager: DialogManager, repo: Repo, **kwargs):
    users = await repo.list_users()
    user_stats = await repo.stats_users()

    return {
        "users": [(user.id, user.id) for user in users],
        "income_today": user_stats.today,
        "income_week": user_stats.week,
        "income_total": user_stats.total,
    }


async def show_user(callback: CallbackQuery, select: Select, dialog_manager: DialogManager, user_id: str):
    await dialog_manager.start(ProfileUserSG.profile, data={"user_id": int(user_id)})


async def search_user(m: Message, text_input: ManagedTextInput, dialog_manager: DialogManager, user_id: int):
    repo: Repo = dialog_manager.middleware_data["repo"]
    user = await repo.get_user(user_id)

    if user is not None:
        await dialog_manager.start(ProfileUserSG.profile, data={"user_id": int(user.id)})


async def change_ban_status(
    callback: CallbackQuery, checkbox: ManagedCheckbox, dialog_manager: DialogManager
):
    repo: Repo = dialog_manager.middleware_data["repo"]
    banned = not checkbox.is_checked()
    user_id = dialog_manager.dialog_data["user_id"]

    await repo.set_user_ban(user_id=user_id, banned=banned)


async def export_users(callback: CallbackQuery, button: Button, dialog_manager: DialogManager):
    repo: Repo = dialog_manager.middleware_data["repo"]
    users = await repo.list_users()

    wb = Workbook()
    ws = wb.active

    ws.row_dimensions[1].height = 30
    for col in "ABCDEFGHI":
        ws.column_dimensions[col].width = 20

    ws["A1"] = "ID"
    ws["B1"] = "Имя"
    ws["C1"] = "Фамилия"
    ws["D1"] = "Никнейм"
    ws["E1"] = "Зарегистрирован"
    ws["F1"] = "Последнее обновление"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for user in users:
        ws.append(
            [
                user.id,
                user.firstname,
                user.lastname,
                user.username,
                user.created_on,
                user.updated_on,
            ]
        )

    # Save the file
    with BytesIO() as table:
        wb.save(table)
        table.seek(0)
        await callback.message.answer_document(
            BufferedInputFile(
                table.read(),
                filename=f"users.{datetime.now().strftime('%Y-%m-%d')}.xlsx",
            )
        )


user_list_dialog = Dialog(
    Window(
        L10NFormat("admin-users-list"),
        ScrollingGroup(
            Select(
                Format("{item[1]}"),
                id="user_list",
                item_id_getter=itemgetter(0),
                items="users",
                on_click=show_user,
            ),
            id="user_list_sg",
            width=1,
            height=10,
            hide_on_single_page=True,
        ),
        TextInput(
            id="fast_search",
            type_factory=process_user_id,
            on_success=search_user,
        ),
        Button(
            L10NFormat("admin-button-export"),
            id="export",
            on_click=export_users,
        ),
        Cancel(L10NFormat("admin-button-back")),
        getter=get_users,
        state=UserSG.lst,
    ),
)


user_profile_dialog = Dialog(
    Window(
        L10NFormat("admin-show-user"),
        Checkbox(
            L10NFormat("admin-show-ban"),
            L10NFormat("admin-show-ban"),
            id="user_ban_ch",
            on_click=change_ban_status,
        ),
        Cancel(L10NFormat("admin-button-back")),
        getter=get_user,
        state=ProfileUserSG.profile,
    ),
)


router.include_routers(
    user_list_dialog,
    user_profile_dialog,
)
